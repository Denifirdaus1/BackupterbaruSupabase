# PythonEngine/main.py

import sys
import os
import logging
import traceback
import time
from file_parser import extract_content, extract_from_image
from analysis_prompt import run_analysis
from output_writer import save_output, save_as_word, save_as_txt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import re

# Setup logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("python_engine_debug.log"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

def get_available_filename(base_path: str, extension: str) -> str:
    logger.debug(f"Mencari nama file yang tersedia untuk {base_path}.{extension}")
    counter = 1
    new_path = f"{base_path}.{extension}"
    while os.path.exists(new_path):
        new_path = f"{base_path} ({counter}).{extension}"
        counter += 1
    logger.debug(f"File output akan disimpan sebagai: {new_path}")
    return new_path

def parse_and_export_excel(output_txt_path: str, output_excel_path: str):
    try:
        logger.info(f"Memulai ekspor Excel dari {output_txt_path}")
        logger.debug(f"Membaca file teks: {output_txt_path}")
        
        with open(output_txt_path, "r", encoding="utf-8") as file:
            content = file.read()
            logger.debug(f"Ukuran konten: {len(content)} karakter")

        # Ambil baris yang merupakan bagian dari tabel markdown
        lines = [line for line in content.splitlines() if '|' in line]
        logger.debug(f"Jumlah baris dengan tabel markdown: {len(lines)}")
        
        if len(lines) < 2:
            logger.warning("Tidak ada tabel markdown yang cukup untuk diekspor ke Excel")
            print("[INFO] Tidak ada tabel markdown untuk diekspor ke Excel.")
            return

        headers = [h.strip() for h in lines[0].strip('|').split('|')]
        rows = [line.strip('|').split('|') for line in lines[2:]]  # Skip garis pemisah
        
        logger.debug(f"Header terdeteksi: {headers}")
        logger.debug(f"Jumlah baris data: {len(rows)}")

        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        logger.debug("Header berhasil ditambahkan ke sheet Excel")

        # Regex untuk mendeteksi formula Excel dalam teks
        formula_regex = r'=\s*([A-Z]+[0-9]+(?:\s*[\+\-\*\/]\s*[A-Z]+[0-9]+)*)'
        sum_regex = r'SUM\s*\(\s*([A-Z]+[0-9]+)\s*:\s*([A-Z]+[0-9]+)\s*\)'
        average_regex = r'AVERAGE\s*\(\s*([A-Z]+[0-9]+)\s*:\s*([A-Z]+[0-9]+)\s*\)'
        
        for i, row in enumerate(rows):
            cells = [cell.strip() for cell in row]
            row_data = []
            
            for cell in cells:
                # Cek apakah cell berisi formula Excel
                if cell.startswith('='):
                    row_data.append(cell)
                    logger.debug(f"Baris {i+1}: Formula langsung: {cell}")
                elif 'SUM(' in cell.upper() or 'JUMLAH(' in cell.upper():
                    sum_match = re.search(sum_regex, cell.upper())
                    if sum_match:
                        formula = f"=SUM({sum_match.group(1)}:{sum_match.group(2)})"
                        row_data.append(formula)
                        logger.debug(f"Baris {i+1}: Mengonversi SUM: {cell} -> {formula}")
                    else:
                        row_data.append(cell)
                        logger.debug(f"Baris {i+1}: Format SUM tidak dikenali: {cell}")
                elif 'AVERAGE(' in cell.upper() or 'RATA-RATA(' in cell.upper():
                    avg_match = re.search(average_regex, cell.upper())
                    if avg_match:
                        formula = f"=AVERAGE({avg_match.group(1)}:{avg_match.group(2)})"
                        row_data.append(formula)
                        logger.debug(f"Baris {i+1}: Mengonversi AVERAGE: {cell} -> {formula}")
                    else:
                        row_data.append(cell)
                        logger.debug(f"Baris {i+1}: Format AVERAGE tidak dikenali: {cell}")
                else:
                    formula_match = re.search(formula_regex, cell)
                    if formula_match:
                        formula = f"={formula_match.group(1)}"
                        row_data.append(formula)
                        logger.debug(f"Baris {i+1}: Mengonversi formula: {cell} -> {formula}")
                    else:
                        row_data.append(cell)
                        logger.debug(f"Baris {i+1}: Data biasa: {cell}")
            
            ws.append(row_data)

        # Auto-fit lebar kolom
        logger.debug("Menyesuaikan lebar kolom...")
        for col in ws.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        logger.info(f"Menyimpan file Excel ke: {output_excel_path}")
        wb.save(output_excel_path)
        print(f"[SUKSES] File Excel hasil perbaikan disimpan ke {output_excel_path}")
        
    except Exception as e:
        logger.error(f"Gagal mengekspor ke Excel: {str(e)}")
        logger.error(traceback.format_exc())
        print(f"[ERROR] Gagal membuat file Excel: {str(e)}")

def main():
    start_time = time.time()
    logger.info("=" * 50)
    logger.info("MEMULAI PROSES PYTHON ENGINE")
    logger.info(f"Arguments: {sys.argv}")
    logger.info("=" * 50)
    
    try:
        if len(sys.argv) < 6:
            error_msg = "Usage: python main.py <file_path or 'none'> <output_txt_path> <user_prompt> <output_format> <mode>"
            logger.error(error_msg)
            print(error_msg)
            return

        file_path = sys.argv[1]
        output_txt_path = sys.argv[2]
        user_prompt = sys.argv[3]
        output_format = sys.argv[4].lower()  # txt, excel, word
        mode = sys.argv[5].lower()           # file, ocr, prompt-only

        logger.info(f"Parameter yang diterima:")
        logger.info(f"  File path: {file_path}")
        logger.info(f"  Output text path: {output_txt_path}")
        logger.info(f"  User prompt: {user_prompt[:50]}... (panjang: {len(user_prompt)})")
        logger.info(f"  Output format: {output_format}")
        logger.info(f"  Mode: {mode}")

        # Ekstrak konten berdasarkan mode
        content = ""
        if mode == "file":
            logger.info("Mode: file - Memproses file input")
            content = extract_content(file_path)
            logger.debug(f"Konten yang diekstrak (100 karakter pertama): {content[:100]}...")
        elif mode == "ocr":
            logger.info("Mode: ocr - Memproses gambar dengan OCR")
            content = extract_from_image(file_path)
            logger.debug(f"Hasil OCR (100 karakter pertama): {content[:100]}...")
        elif mode == "prompt-only":
            logger.info("Mode: prompt-only - Tanpa file input")
            content = ""
        else:
            error_msg = f"[ERROR] Mode tidak dikenali: {mode}"
            logger.error(error_msg)
            print(error_msg)
            return

        logger.info("Memulai analisis dengan Gemini API...")
        result_text = run_analysis(content, user_prompt, output_format)
        logger.debug(f"Hasil analisis Gemini (100 karakter pertama): {result_text[:100]}...")
        logger.debug(f"Panjang hasil analisis: {len(result_text)} karakter")

        logger.info(f"Menyimpan hasil sementara ke {output_txt_path}")
        save_output(result_text, output_txt_path)

        # Tentukan path aman untuk output
        base_path = output_txt_path.rsplit(".", 1)[0]
        logger.info(f"Base path untuk output: {base_path}")

        if output_format == "txt":
            logger.info("Format output: TXT")
            final_txt_path = get_available_filename(base_path + "_final", "txt")
            save_as_txt(result_text, final_txt_path)
            logger.info(f"File TXT final disimpan di: {final_txt_path}")

        elif output_format == "excel":
            logger.info("Format output: EXCEL")
            safe_excel_path = get_available_filename(base_path + "_parsed", "xlsx")
            logger.info(f"Memulai ekspor ke Excel: {safe_excel_path}")
            parse_and_export_excel(output_txt_path, safe_excel_path)

        elif output_format == "word":
            logger.info("Format output: WORD")
            safe_docx_path = get_available_filename(base_path + "_output", "docx")
            logger.info(f"Menyimpan ke dokumen Word: {safe_docx_path}")
            save_as_word(result_text, safe_docx_path)
            logger.info("Dokumen Word berhasil dibuat")

        else:
            error_msg = f"[ERROR] Format output tidak dikenali: {output_format}"
            logger.error(error_msg)
            print(error_msg)
            return

        logger.info("Proses berhasil diselesaikan")
        print("OK")
        
    except Exception as e:
        logger.error(f"ERROR TIDAK TERDUGA: {str(e)}")
        logger.error(traceback.format_exc())
        print(f"[FATAL ERROR] {str(e)}")
    finally:
        duration = time.time() - start_time
        logger.info(f"Proses selesai dalam {duration:.2f} detik")
        logger.info("=" * 50)

if __name__ == "__main__":
    main()
